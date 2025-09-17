# DGII Report Routes - Dominican Republic Tax Compliance
from flask import Blueprint, request, jsonify, send_file, current_app
from flask import render_template, session, redirect, url_for, flash
import csv
import io
import tempfile
import os
from datetime import datetime, timedelta
import calendar
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from models import db, User, Sale, Purchase, Product, Supplier, NCFSequence
from utils import get_company_info_for_receipt

bp = Blueprint('dgii', __name__, url_prefix='/dgii')


def require_admin():
    """Check if user is authenticated and is an admin"""
    # Check if this is an AJAX/JSON request
    is_json_request = request.is_json or 'application/json' in request.headers.get('Content-Type', '')
    
    if 'user_id' not in session:
        if is_json_request:
            return None  # Will be handled by calling function
        return redirect(url_for('auth.login'))
    
    user = User.query.get(session['user_id'])
    if not user or not user.active:
        session.clear()
        if is_json_request:
            return None  # Will be handled by calling function
        return redirect(url_for('auth.login'))
    
    if user.role.value != 'ADMINISTRADOR':
        if is_json_request:
            return None  # Will be handled by calling function
        flash('Acceso denegado. Se requieren permisos de administrador.', 'error')
        return redirect(url_for('admin.dashboard'))
    
    return user

# DGII 606 Layout - Compras (Purchases)
# Layout for purchases from suppliers with tax details
DGII_606_HEADERS = [
    'RNC_CEDULA',           # RNC/Cédula del Proveedor
    'TIPO_IDENTIFICACION',  # Tipo de identificación (1=RNC, 2=Cédula)
    'NUMERO_COMPROBANTE_FISCAL',  # NCF del Comprobante
    'NUMERO_COMPROBANTE_MODIFICADO',  # NCF del comprobante que modifica (si aplica)
    'FECHA_COMPROBANTE',    # Fecha del comprobante (YYYYMMDD)
    'FECHA_PAGO',          # Fecha de pago (YYYYMMDD)
    'MONTO_FACTURADO',     # Monto total facturado
    'ITBIS_FACTURADO',     # ITBIS facturado
    'ITBIS_RETENIDO',      # ITBIS retenido por el comprador
    'ITBIS_SUJETO_PROPORCION',  # ITBIS sujeto a proporcionalidad
    'ITBIS_LLEVADO_COSTO',  # ITBIS llevado al costo
    'ITBIS_POR_ADELANTAR',  # ITBIS por adelantar
    'ITBIS_PERCIBIDO_COMPRAS',  # ITBIS percibido en compras
    'TIPO_RETENCION_ISR',   # Tipo de retención ISR
    'MONTO_RETENCION_ISR',  # Monto retención ISR
    'TIPO_ANULACION'        # Tipo de anulación
]

# DGII 607 Layout - Ventas (Sales)  
# Layout for sales to customers with tax details
DGII_607_HEADERS = [
    'RNC_CEDULA',           # RNC/Cédula del Cliente
    'TIPO_IDENTIFICACION',  # Tipo de identificación (1=RNC, 2=Cédula)
    'NUMERO_COMPROBANTE_FISCAL',  # NCF del Comprobante
    'NUMERO_COMPROBANTE_MODIFICADO',  # NCF del comprobante que modifica (si aplica)
    'FECHA_COMPROBANTE',    # Fecha del comprobante (YYYYMMDD)
    'MONTO_FACTURADO',     # Monto total facturado
    'ITBIS_FACTURADO',     # ITBIS facturado
    'ITBIS_RETENIDO',      # ITBIS retenido al vendedor
    'ITBIS_PERCIBIDO',     # ITBIS percibido
    'RETENCION_RENTA',     # Retención de renta
    'ISC',                 # Impuesto Selectivo al Consumo
    'OTROS_IMPUESTOS',     # Otros impuestos/tasas
    'MONTO_PROPINA_LEGAL', # Monto propina legal
    'TIPO_ANULACION'       # Tipo de anulación
]


@bp.route('/')
def reports():
    """DGII Reports dashboard"""
    user = require_admin()
    if not isinstance(user, User):
        return redirect(url_for('auth.login'))
    
    # Get available months for dropdown
    current_date = datetime.now()
    available_months = []
    
    # Last 12 months
    for i in range(12):
        month_date = current_date - timedelta(days=30*i)
        month_info = {
            'year': month_date.year,
            'month': month_date.month,
            'name': calendar.month_name[month_date.month],
            'value': f"{month_date.year}-{month_date.month:02d}"
        }
        available_months.append(month_info)
    
    # Get summary stats for current month
    current_month_start = current_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    if current_date.month == 12:
        next_month_start = current_date.replace(year=current_date.year + 1, month=1, day=1)
    else:
        next_month_start = current_date.replace(month=current_date.month + 1, day=1)
    
    # Current month sales count
    sales_count = Sale.query.filter(
        Sale.created_at >= current_month_start,
        Sale.created_at < next_month_start
    ).count()
    
    # Current month purchases count  
    purchases_count = Purchase.query.filter(
        Purchase.created_at >= current_month_start,
        Purchase.created_at < next_month_start
    ).count()
    
    return render_template('dgii/reports.html',
                         available_months=available_months,
                         current_month=current_date.strftime("%B %Y"),
                         sales_count=sales_count,
                         purchases_count=purchases_count)


@bp.route('/export/606', methods=['POST'])
def export_606():
    """Export DGII 606 (Purchases) CSV"""
    user = require_admin()
    if not isinstance(user, User):
        return jsonify({'error': 'No autorizado'}), 401
    
    # Accept multiple input formats (JSON, form data, or query parameters)
    data = request.get_json(silent=True) or request.form or request.args
    if not data:
        return jsonify({'error': 'Datos no proporcionados'}), 400
    
    # Support both individual year/month and period format (YYYY-MM)
    if data.get('period'):
        try:
            year, month = map(int, data['period'].split('-')[:2])
        except (ValueError, AttributeError):
            return jsonify({'error': 'Formato de período inválido. Use YYYY-MM'}), 400
    else:
        year = data.get('year')
        month = data.get('month')
        
        if not year or not month:
            return jsonify({'error': 'Año y mes son requeridos'}), 400
    
    try:
        year = int(year)
        month = int(month)
        
        # Validate month range
        if not (1 <= month <= 12):
            return jsonify({'error': 'Mes debe estar entre 1 y 12'}), 400
        
        # Date range for the month
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 1, 1)
        
        # Get purchases for the period
        purchases = Purchase.query.filter(
            Purchase.created_at >= start_date,
            Purchase.created_at < end_date
        ).join(Supplier).all()
        
        # Generate CSV
        output = io.StringIO()
        writer = csv.writer(output, delimiter='|')
        
        # Write header
        writer.writerow(DGII_606_HEADERS)
        
        # Get company information for default RNC
        company_info = get_company_info_for_receipt()
        company_rnc = company_info.get('rnc', '').replace('-', '')  # Remove dashes for DGII format
        
        # Write purchase records
        for purchase in purchases:
            supplier = purchase.supplier
            
            # Determine identification type
            if supplier.rnc and len(supplier.rnc) == 9:
                tipo_id = '1'  # RNC
                rnc_cedula = supplier.rnc
            elif supplier.rnc and len(supplier.rnc) == 11:
                tipo_id = '2'  # Cédula  
                rnc_cedula = supplier.rnc
            else:
                # Use company RNC if available, otherwise default
                tipo_id = '1'  # Default to RNC
                if company_rnc and len(company_rnc) >= 9:
                    rnc_cedula = company_rnc[:9] if len(company_rnc) > 9 else company_rnc.ljust(9, '0')
                else:
                    rnc_cedula = '000000000'
            
            # Format dates as YYYYMMDD
            fecha_comprobante = purchase.created_at.strftime('%Y%m%d')
            fecha_pago = purchase.created_at.strftime('%Y%m%d')  # Assume same date
            
            # Calculate amounts
            monto_facturado = purchase.total_amount
            itbis_facturado = purchase.tax_amount or 0
            
            # Write row
            row = [
                rnc_cedula,                    # RNC_CEDULA
                tipo_id,                       # TIPO_IDENTIFICACION
                purchase.ncf_supplier or '',   # NUMERO_COMPROBANTE_FISCAL
                '',                           # NUMERO_COMPROBANTE_MODIFICADO
                fecha_comprobante,            # FECHA_COMPROBANTE
                fecha_pago,                   # FECHA_PAGO
                f"{monto_facturado:.2f}",     # MONTO_FACTURADO
                f"{itbis_facturado:.2f}",     # ITBIS_FACTURADO
                '0.00',                       # ITBIS_RETENIDO
                '0.00',                       # ITBIS_SUJETO_PROPORCION
                '0.00',                       # ITBIS_LLEVADO_COSTO
                '0.00',                       # ITBIS_POR_ADELANTAR
                '0.00',                       # ITBIS_PERCIBIDO_COMPRAS
                '',                           # TIPO_RETENCION_ISR
                '0.00',                       # MONTO_RETENCION_ISR
                ''                            # TIPO_ANULACION
            ]
            writer.writerow(row)
        
        # Create temporary file
        csv_content = output.getvalue()
        output.close()
        
        # Generate filename
        filename = f"606_{year}_{month:02d}.csv"
        
        # Return file
        return jsonify({
            'success': True,
            'filename': filename,
            'records': len(purchases),
            'csv_data': csv_content,
            'message': f'Reporte 606 generado: {len(purchases)} compras para {calendar.month_name[month]} {year}'
        })
        
    except Exception as e:
        return jsonify({'error': f'Error generando reporte 606: {str(e)}'}), 400


@bp.route('/export/607', methods=['POST'])
def export_607():
    """Export DGII 607 (Sales) CSV"""
    user = require_admin()
    
    if not isinstance(user, User):
        return jsonify({'error': 'No autorizado'}), 401
    
    # Accept multiple input formats (JSON, form data, or query parameters)
    data = request.get_json(silent=True) or request.form or request.args
    
    if not data:
        return jsonify({'error': 'Datos no proporcionados'}), 400
    
    # Support both individual year/month and period format (YYYY-MM)
    if data.get('period'):
        try:
            year, month = map(int, data['period'].split('-')[:2])
        except (ValueError, AttributeError):
            current_app.logger.error("[DEBUG DGII 607] Invalid period format")
            return jsonify({'error': 'Formato de período inválido. Use YYYY-MM'}), 400
    else:
        year = data.get('year')
        month = data.get('month')
        
        if not year or not month:
            current_app.logger.error("[DEBUG DGII 607] Missing year or month")
            return jsonify({'error': 'Año y mes son requeridos'}), 400
    
    current_app.logger.debug(f"[DEBUG DGII 607] Year: {year}, Month: {month}")
    
    try:
        year = int(year)
        month = int(month)
        
        # Validate month range
        if not (1 <= month <= 12):
            current_app.logger.error("[DEBUG DGII 607] Invalid month range")
            return jsonify({'error': 'Mes debe estar entre 1 y 12'}), 400
            
        current_app.logger.debug(f"[DEBUG DGII 607] Converted Year: {year}, Month: {month}")
        
        # Date range for the month
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 1, 1)
        
        # Get sales for the period (only completed sales)
        current_app.logger.debug(f"[DEBUG DGII 607] Querying sales from {start_date} to {end_date}")
        sales = Sale.query.filter(
            Sale.created_at >= start_date,
            Sale.created_at < end_date,
            Sale.status == 'completed'
        ).all()
        
        current_app.logger.info(f"[DEBUG DGII 607] Found {len(sales)} sales")
        
        # Get company information for default RNC
        company_info = get_company_info_for_receipt()
        company_rnc = company_info.get('rnc', '').replace('-', '')  # Remove dashes for DGII format
        
        # Generate CSV
        output = io.StringIO()
        writer = csv.writer(output, delimiter='|')
        
        # Write header
        writer.writerow(DGII_607_HEADERS)
        current_app.logger.debug("[DEBUG DGII 607] CSV header written")
        
        # Write sales records
        for sale in sales:
            current_app.logger.debug(f"[DEBUG DGII 607] Processing sale {sale.id}")
            # Determine identification type based on customer RNC/Cédula
            if sale.customer_rnc and len(sale.customer_rnc) == 9:
                tipo_id = '1'  # RNC
                rnc_cedula = sale.customer_rnc
            elif sale.customer_rnc and len(sale.customer_rnc) == 11:
                tipo_id = '2'  # Cédula
                rnc_cedula = sale.customer_rnc
            else:
                # Use company RNC if available for general public sales
                if company_rnc and len(company_rnc) >= 9:
                    tipo_id = '1'  # Use RNC
                    rnc_cedula = company_rnc[:9] if len(company_rnc) > 9 else company_rnc.ljust(9, '0')
                else:
                    # Fallback to cédula format for general public
                    tipo_id = '2'  # Default to Cédula for general public
                    rnc_cedula = '00000000000'  # Default for general public
            
            # Format date as YYYYMMDD
            fecha_comprobante = sale.created_at.strftime('%Y%m%d')
            
            # Calculate amounts
            monto_facturado = sale.total
            itbis_facturado = sale.tax_amount or 0
            
            # Write row
            row = [
                rnc_cedula,                    # RNC_CEDULA
                tipo_id,                       # TIPO_IDENTIFICACION
                sale.ncf,                      # NUMERO_COMPROBANTE_FISCAL
                '',                            # NUMERO_COMPROBANTE_MODIFICADO
                fecha_comprobante,             # FECHA_COMPROBANTE
                f"{monto_facturado:.2f}",      # MONTO_FACTURADO
                f"{itbis_facturado:.2f}",      # ITBIS_FACTURADO
                '0.00',                        # ITBIS_RETENIDO
                '0.00',                        # ITBIS_PERCIBIDO
                '0.00',                        # RETENCION_RENTA
                '0.00',                        # ISC
                '0.00',                        # OTROS_IMPUESTOS
                '0.00',                        # MONTO_PROPINA_LEGAL
                ''                             # TIPO_ANULACION
            ]
            writer.writerow(row)
        
        # Create CSV content
        csv_content = output.getvalue()
        output.close()
        
        # Generate filename
        filename = f"607_{year}_{month:02d}.csv"
        
        current_app.logger.info("[DEBUG DGII 607] CSV processing completed successfully")
        
        # Return file data
        response_data = {
            'success': True,
            'filename': filename,
            'records': len(sales),
            'csv_data': csv_content,
            'message': f'Reporte 607 generado: {len(sales)} ventas para {calendar.month_name[month]} {year}'
        }
        current_app.logger.info(f"[DEBUG DGII 607] Response data prepared: {response_data['message']}")
        return jsonify(response_data)
        
    except Exception as e:
        current_app.logger.exception(f"[DEBUG DGII 607] Exception occurred: {str(e)}")
        return jsonify({'error': f'Error generando reporte 607: {str(e)}'}), 400


@bp.route('/preview/606/<int:year>/<int:month>')
def preview_606(year, month):
    """Preview DGII 606 data before export"""
    user = require_admin()
    if not isinstance(user, User):
        return redirect(url_for('auth.login'))
    
    try:
        # Date range for the month
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 1, 1)
        
        # Get purchases for the period
        purchases = Purchase.query.filter(
            Purchase.created_at >= start_date,
            Purchase.created_at < end_date
        ).join(Supplier).all()
        
        return render_template('dgii/preview_606.html',
                             purchases=purchases,
                             year=year,
                             month=month,
                             month_name=calendar.month_name[month])
        
    except Exception as e:
        flash(f'Error cargando vista previa: {str(e)}', 'error')
        return redirect(url_for('dgii.reports'))


@bp.route('/preview/607/<int:year>/<int:month>')
def preview_607(year, month):
    """Preview DGII 607 data before export"""
    user = require_admin()
    if not isinstance(user, User):
        return redirect(url_for('auth.login'))
    
    try:
        # Date range for the month
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 1, 1)
        
        # Get sales for the period (only completed sales)
        sales = Sale.query.filter(
            Sale.created_at >= start_date,
            Sale.created_at < end_date,
            Sale.status == 'completed'
        ).all()
        
        return render_template('dgii/preview_607.html',
                             sales=sales,
                             year=year,
                             month=month,
                             month_name=calendar.month_name[month])
        
    except Exception as e:
        flash(f'Error cargando vista previa: {str(e)}', 'error')
        return redirect(url_for('dgii.reports'))


# EXPORT TO EXCEL FUNCTIONALITY
@bp.route('/export/606/excel', methods=['POST'])
def export_606_excel():
    """Export DGII 606 (Purchases) to Excel format"""
    user = require_admin()
    if not isinstance(user, User):
        return jsonify({'error': 'No autorizado'}), 401
    
    # Accept multiple input formats (JSON, form data, or query parameters)
    data = request.get_json(silent=True) or request.form or request.args
    if not data:
        return jsonify({'error': 'Datos no proporcionados'}), 400
    
    # Support both individual year/month and period format (YYYY-MM)
    if data.get('period'):
        try:
            year, month = map(int, data['period'].split('-')[:2])
        except (ValueError, AttributeError):
            return jsonify({'error': 'Formato de período inválido. Use YYYY-MM'}), 400
    else:
        year = data.get('year')
        month = data.get('month')
        
        if not year or not month:
            return jsonify({'error': 'Año y mes son requeridos'}), 400
    
    try:
        year = int(year)
        month = int(month)
        
        # Validate month range
        if not (1 <= month <= 12):
            return jsonify({'error': 'Mes debe estar entre 1 y 12'}), 400
            
        current_app.logger.debug(f'DGII 606 Excel export started: {year}-{month:02d}')
        
        # Get purchases for the period
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 1, 1)
        
        purchases = Purchase.query.filter(
            Purchase.created_at >= start_date,
            Purchase.created_at < end_date
        ).join(Supplier).all()
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet('DGII 606')
            wb.active = ws
        ws.title = f"DGII 606 - {calendar.month_name[month]} {year}"
        current_app.logger.debug(f'Excel workbook created with {len(purchases)} purchases')
        
        # Add title
        ws.merge_cells('A1:P1')
        title_cell = ws['A1']
        title_cell.value = f"Reporte DGII 606 - Compras - {calendar.month_name[month]} {year}"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        title_cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        title_cell.font = Font(color='FFFFFF', size=16, bold=True)
        
        # Add headers
        for col, header in enumerate(DGII_606_HEADERS, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        
        # Add data rows
        row_num = 4
        for purchase in purchases:
            supplier = purchase.supplier
            
            # Determine identification type
            if supplier.rnc and len(supplier.rnc) == 9:
                tipo_id = '1'  # RNC
                rnc_cedula = supplier.rnc
            elif supplier.rnc and len(supplier.rnc) == 11:
                tipo_id = '2'  # Cédula  
                rnc_cedula = supplier.rnc
            else:
                tipo_id = '1'  # Default to RNC
                rnc_cedula = supplier.rnc or '000000000'
            
            # Format dates as YYYYMMDD
            fecha_comprobante = purchase.created_at.strftime('%Y%m%d')
            fecha_pago = purchase.created_at.strftime('%Y%m%d')
            
            # Calculate amounts
            monto_facturado = purchase.total_amount
            itbis_facturado = purchase.tax_amount or 0
            
            # Data row
            row_data = [
                rnc_cedula, tipo_id, purchase.ncf_supplier or '', '',
                fecha_comprobante, fecha_pago, f"{monto_facturado:.2f}",
                f"{itbis_facturado:.2f}", '0.00', '0.00', '0.00', '0.00',
                '0.00', '', '0.00', ''
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col, value=value)
            
            row_num += 1
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        filename = f"606_{year}_{month:02d}.xlsx"
        
        return send_file(temp_file.name, as_attachment=True, download_name=filename,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
    except Exception as e:
        current_app.logger.exception(f'Error generating Excel 606: {str(e)}')
        return jsonify({'error': f'Error generando Excel 606: {str(e)}'}), 400


@bp.route('/export/607/excel', methods=['POST'])
def export_607_excel():
    """Export DGII 607 (Sales) to Excel format"""
    user = require_admin()
    if not isinstance(user, User):
        return jsonify({'error': 'No autorizado'}), 401
    
    # Accept multiple input formats (JSON, form data, or query parameters)
    data = request.get_json(silent=True) or request.form or request.args
    if not data:
        return jsonify({'error': 'Datos no proporcionados'}), 400
    
    # Support both individual year/month and period format (YYYY-MM)
    if data.get('period'):
        try:
            year, month = map(int, data['period'].split('-')[:2])
        except (ValueError, AttributeError):
            return jsonify({'error': 'Formato de período inválido. Use YYYY-MM'}), 400
    else:
        year = data.get('year')
        month = data.get('month')
        
        if not year or not month:
            return jsonify({'error': 'Año y mes son requeridos'}), 400
    
    try:
        year, month = int(year), int(month)
        
        # Validate month range
        if not (1 <= month <= 12):
            return jsonify({'error': 'Mes debe estar entre 1 y 12'}), 400
            
        current_app.logger.debug(f'DGII 607 Excel export started: {year}-{month:02d}')
        start_date = datetime(year, month, 1)
        end_date = datetime(year + 1, 1, 1) if month == 12 else datetime(year, month + 1, 1)
        
        sales = Sale.query.filter(
            Sale.created_at >= start_date, Sale.created_at < end_date,
            Sale.status == 'completed'
        ).all()
        
        # Create Excel workbook
        wb = openpyxl.Workbook()  # Use default mode, not write_only
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet('DGII 607')
        ws.title = f"DGII 607 - {calendar.month_name[month]} {year}"
        
        # Add title and headers
        ws.merge_cells('A1:N1')
        ws['A1'] = f"Reporte DGII 607 - Ventas - {calendar.month_name[month]} {year}"
        ws['A1'].font = Font(color='FFFFFF', size=16, bold=True)
        ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        for col, header in enumerate(DGII_607_HEADERS, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        
        # Get company information for default RNC
        company_info = get_company_info_for_receipt()
        company_rnc = company_info.get('rnc', '').replace('-', '')  # Remove dashes for DGII format
        
        # Add data
        for row_num, sale in enumerate(sales, 4):
            if sale.customer_rnc and len(sale.customer_rnc) == 9:
                tipo_id, rnc_cedula = '1', sale.customer_rnc
            elif sale.customer_rnc and len(sale.customer_rnc) == 11:
                tipo_id, rnc_cedula = '2', sale.customer_rnc
            else:
                # Use company RNC if available for general public sales
                if company_rnc and len(company_rnc) >= 9:
                    tipo_id = '1'  # Use RNC
                    rnc_cedula = company_rnc[:9] if len(company_rnc) > 9 else company_rnc.ljust(9, '0')
                else:
                    # Fallback to cédula format for general public
                    tipo_id, rnc_cedula = '2', '00000000000'
            
            row_data = [
                rnc_cedula, tipo_id, sale.ncf, '',
                sale.created_at.strftime('%Y%m%d'), f"{sale.total:.2f}",
                f"{sale.tax_amount or 0:.2f}", '0.00', '0.00', '0.00', '0.00', '0.00', '0.00', ''
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col, value=value)
        
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        return send_file(temp_file.name, as_attachment=True, 
                        download_name=f"607_{year}_{month:02d}.xlsx",
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
    except Exception as e:
        return jsonify({'error': f'Error generando Excel 607: {str(e)}'}), 400


@bp.route('/export/607/pdf', methods=['POST'])
def export_607_pdf():
    """Export DGII 607 (Sales) to PDF format"""
    try:
        user = require_admin()
        
        if not isinstance(user, User):
            return jsonify({'error': 'No autorizado'}), 401
        
        # Accept multiple input formats (JSON, form data, or query parameters)
        data = request.get_json(silent=True) or request.form or request.args
        if not data:
            return jsonify({'error': 'Datos no proporcionados'}), 400
        
        # Support both individual year/month and period format (YYYY-MM)
        if data.get('period'):
            try:
                year, month = map(int, data['period'].split('-')[:2])
            except (ValueError, AttributeError):
                return jsonify({'error': 'Formato de período inválido. Use YYYY-MM'}), 400
        else:
            year = data.get('year')
            month = data.get('month')
            
            if not year or not month:
                return jsonify({'error': 'Año y mes son requeridos'}), 400
        
        year = int(year)
        month = int(month)
        
        # Validate month range
        if not (1 <= month <= 12):
            return jsonify({'error': 'Mes debe estar entre 1 y 12'}), 400
            
        current_app.logger.debug(f'DGII 607 PDF export started: {year}-{month:02d}')
        
        # Date range for the month
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 1, 1)
        
        # Get sales for the period
        sales = Sale.query.filter(
            Sale.created_at >= start_date,
            Sale.created_at < end_date,
            Sale.status == 'completed'
        ).all()
        
        # Get company information
        company_info = get_company_info_for_receipt()
        company_rnc = company_info.get('rnc', '').replace('-', '')  # Remove dashes for DGII format
        
        # Create professional PDF using SimpleDocTemplate
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        doc = SimpleDocTemplate(temp_file.name, pagesize=A4)
        story = []
        
        # Get styles
        styles = getSampleStyleSheet()
        
        # Create custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.darkblue,
            alignment=1,  # Center
            spaceAfter=20
        )
        
        company_style = ParagraphStyle(
            'CompanyInfo',
            parent=styles['Normal'],
            fontSize=10,
            alignment=1,  # Center
            spaceAfter=15
        )
        
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.darkblue,
            alignment=1,  # Center
            spaceAfter=15
        )
        
        # Add title
        title = Paragraph(f"REPORTE DGII 607 - VENTAS", title_style)
        story.append(title)
        
        # Add company information
        if company_info.get('name'):
            company_name = Paragraph(f"<b>{company_info['name']}</b>", company_style)
            story.append(company_name)
        
        if company_info.get('rnc'):
            company_rnc_display = Paragraph(f"RNC: {company_info['rnc']}", company_style)
            story.append(company_rnc_display)
        
        if company_info.get('address'):
            company_address = Paragraph(f"Dirección: {company_info['address']}", company_style)
            story.append(company_address)
        
        # Add period and summary
        period_info = Paragraph(f"<b>Período: {calendar.month_name[month]} {year}</b>", subtitle_style)
        story.append(period_info)
        
        summary_info = Paragraph(f"<b>Total de transacciones: {len(sales)}</b><br/>Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}", company_style)
        story.append(summary_info)
        
        if sales:
            # Calculate totals
            total_monto = sum(sale.total for sale in sales)
            total_itbis = sum(sale.tax_amount or 0 for sale in sales)
            
            # Create data table
            table_data = [
                ['RNC/Cédula', 'Tipo ID', 'NCF', 'Fecha', 'Monto Facturado', 'ITBIS', 'Cliente', 'Observaciones']
            ]
            
            for sale in sales:
                # Determine identification type
                if sale.customer_rnc and len(sale.customer_rnc) == 9:
                    tipo_id = '1'  # RNC
                    rnc_cedula = sale.customer_rnc
                elif sale.customer_rnc and len(sale.customer_rnc) == 11:
                    tipo_id = '2'  # Cédula
                    rnc_cedula = sale.customer_rnc
                else:
                    # Use company RNC if available for general public sales
                    if company_rnc and len(company_rnc) >= 9:
                        tipo_id = '1'  # Use RNC
                        rnc_cedula = company_rnc[:9] if len(company_rnc) > 9 else company_rnc.ljust(9, '0')
                        customer_display = 'Público General'
                    else:
                        tipo_id = '2'  # Default to Cédula for general public
                        rnc_cedula = '00000000000'
                        customer_display = 'Público General'
                
                # Format customer display
                if sale.customer_name:
                    customer_display = sale.customer_name[:20] + ('...' if len(sale.customer_name) > 20 else '')
                elif hasattr(sale, 'customer_rnc') and sale.customer_rnc:
                    customer_display = 'Cliente con RNC'
                else:
                    customer_display = 'Público General'
                
                table_data.append([
                    rnc_cedula,
                    tipo_id,
                    sale.ncf or 'N/A',
                    sale.created_at.strftime('%d/%m/%Y'),
                    f"${sale.total:,.2f}",
                    f"${sale.tax_amount or 0:,.2f}",
                    customer_display,
                    'Venta completada'
                ])
            
            # Add totals row
            table_data.append([
                '', '', '', '<b>TOTALES:</b>', f"<b>${total_monto:,.2f}</b>", f"<b>${total_itbis:,.2f}</b>", '', ''
            ])
            
            # Create table
            table = Table(table_data, colWidths=[80, 40, 80, 60, 80, 60, 80, 80])
            table.setStyle(TableStyle([
                # Header row
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Data rows
                ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -2), colors.black),
                ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -2), 7),
                ('GRID', (0, 0), (-1, -2), 1, colors.black),
                
                # Totals row
                ('BACKGROUND', (0, -1), (-1, -1), colors.grey),
                ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 8),
                ('GRID', (0, -1), (-1, -1), 1, colors.black),
            ]))
            
            story.append(table)
            
            # Add summary paragraph
            summary_text = f"""
            <b>RESUMEN DEL REPORTE:</b><br/>
            • Total de transacciones procesadas: {len(sales)}<br/>
            • Monto total facturado: ${total_monto:,.2f}<br/>
            • ITBIS total facturado: ${total_itbis:,.2f}<br/>
            • Período: {calendar.month_name[month]} {year}<br/>
            • Generado el: {datetime.now().strftime('%d de %B de %Y a las %H:%M')}
            """
            summary_para = Paragraph(summary_text, styles['Normal'])
            story.append(summary_para)
        else:
            no_data = Paragraph("No se encontraron ventas para el período seleccionado.", styles['Normal'])
            story.append(no_data)
        
        # Add footer
        footer_text = f"""
        <br/><br/>
        <i>Este reporte ha sido generado automáticamente por el sistema de gestión fiscal.<br/>
        Para consultas o aclaraciones, contacte al departamento de contabilidad.</i>
        """
        footer = Paragraph(footer_text, styles['Normal'])
        story.append(footer)
        
        # Build PDF
        doc.build(story)
        temp_file.close()
        
        return send_file(temp_file.name, as_attachment=True, 
                        download_name=f"607_{year}_{month:02d}.pdf",
                        mimetype='application/pdf')
    
    except Exception as e:
        current_app.logger.error(f"Error generando PDF 607: {str(e)}")
        return jsonify({'error': f'Error generando PDF 607: {str(e)}'}), 400


@bp.route('/export/607/txt', methods=['POST'])
def export_607_txt():
    """Export DGII 607 (Sales) to TXT format (DGII compliant)"""
    try:
        current_app.logger.info("Starting TXT export request")
        
        user = require_admin()
        current_app.logger.info(f"User check result: {type(user)} - {user}")
        
        if not isinstance(user, User):
            current_app.logger.error(f"User auth failed: {type(user)}")
            return jsonify({'error': 'No autorizado'}), 401
        
        # Accept multiple input formats (JSON, form data, or query parameters)
        data = request.get_json(silent=True) or request.form or request.args
        if not data:
            return jsonify({'error': 'Datos no proporcionados'}), 400
        
        # Support both individual year/month and period format (YYYY-MM)
        if data.get('period'):
            try:
                year, month = map(int, data['period'].split('-')[:2])
            except (ValueError, AttributeError):
                return jsonify({'error': 'Formato de período inválido. Use YYYY-MM'}), 400
        else:
            year = data.get('year')
            month = data.get('month')
            
            if not year or not month:
                return jsonify({'error': 'Año y mes son requeridos'}), 400
        
        year = int(year)
        month = int(month)
        
        # Validate month range
        if not (1 <= month <= 12):
            return jsonify({'error': 'Mes debe estar entre 1 y 12'}), 400
            
        current_app.logger.debug(f'DGII 607 TXT export started: {year}-{month:02d}')
        
        # Date range for the month
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 1, 1)
        
        # Get sales for the period
        sales = Sale.query.filter(
            Sale.created_at >= start_date,
            Sale.created_at < end_date,
            Sale.status == 'completed'
        ).all()
        
        # Get company information for default RNC
        company_info = get_company_info_for_receipt()
        company_rnc = company_info.get('rnc', '').replace('-', '')  # Remove dashes for DGII format
        
        # Generate TXT content with DGII format (pipe-delimited)
        lines = []
        
        for sale in sales:
            # Determine identification type
            if sale.customer_rnc and len(sale.customer_rnc) == 9:
                tipo_id = '1'  # RNC
                rnc_cedula = sale.customer_rnc
            elif sale.customer_rnc and len(sale.customer_rnc) == 11:
                tipo_id = '2'  # Cédula
                rnc_cedula = sale.customer_rnc
            else:
                # Use company RNC if available for general public sales
                if company_rnc and len(company_rnc) >= 9:
                    tipo_id = '1'  # Use RNC
                    rnc_cedula = company_rnc[:9] if len(company_rnc) > 9 else company_rnc.ljust(9, '0')
                else:
                    # Fallback to cédula format for general public
                    tipo_id = '2'  # Default to Cédula for general public
                    rnc_cedula = '00000000000'
                
            fecha_comprobante = sale.created_at.strftime('%Y%m%d')
            monto_facturado = f"{sale.total:.2f}"
            itbis_facturado = f"{sale.tax_amount or 0:.2f}"
            
            # Format: All fields separated by pipe (|) as per DGII specification
            line = f"{rnc_cedula}|{tipo_id}|{sale.ncf or ''}||{fecha_comprobante}|{monto_facturado}|{itbis_facturado}|0.00|0.00|0.00|0.00|0.00|0.00|"
            lines.append(line)
        
        txt_content = '\n'.join(lines)
        filename = f"607_{year}_{month:02d}.txt"
        
        # Return as JSON (like CSV export)
        return jsonify({
            'success': True,
            'filename': filename,
            'records': len(sales),
            'txt_data': txt_content,
            'message': f'Reporte 607 TXT generado: {len(sales)} ventas para {calendar.month_name[month]} {year}'
        })
        
    except Exception as e:
        current_app.logger.error(f"Error generando TXT 607: {str(e)}")
        return jsonify({'error': f'Error generando TXT 607: {str(e)}'}), 400